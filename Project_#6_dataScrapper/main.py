import xlrd
import openpyxl
import os
import math
# Get rid of nonrelevant warning
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

#Methods
def is_float(string):
    try:
        float(string)
        return True
    except ValueError:
        return False



# Note: .xls files using "xlrd" package use Zero-based[0] array indexing
# while .xlsx files using "openpyxl" package use One-based[1] array indexing
# A parenthesis with "1" or "0" will denote index base when indexed

#Input Data Path declaration
mySourcePath = os.path.join(os.path.abspath(os.path.dirname(__file__)), "input", "TDP.xls")
#Input Format Path declaration
myFormatPath = os.path.join(os.path.abspath(os.path.dirname(__file__)), "input", "dataSheet.xlsx")
#Output Path declaration
myDestPath = os.path.join(os.path.abspath(os.path.dirname(__file__)), "output", "destinyDataSheet.xlsx")

# To open the source workbook, "wb_Source" object is created
wbSource = xlrd.open_workbook(mySourcePath)
# To open the format workbook , "wbFormat" workbook object is created
wbFormat = openpyxl.load_workbook(myFormatPath)

# (0)Get source workbook sheet object from index
sourceActSheet = wbSource.sheet_by_index(1)
#print("Cell D30 is {0}".format(sourceActSheet.cell_value(rowx=1, colx=0)))
# (1)Get destination workbook sheet object from name
destinationActSheet = wbFormat["equipmentData"]

# (1)(0)Populate row from header values
destinationActSheet.cell(row=2, column=1, value=sourceActSheet.cell_value(rowx=1, colx=2))#Equipment
destinationActSheet.cell(row=2, column=2, value=sourceActSheet.cell_value(rowx=1, colx=5))#Source
destinationActSheet.cell(row=2, column=3, value=sourceActSheet.cell_value(rowx=3, colx=2))#Location
destinationActSheet.cell(row=2, column=4, value="")#Room
destinationActSheet.cell(row=2, column=5, value=sourceActSheet.cell_value(rowx=2, colx=6))#Voltage LL
destinationActSheet.cell(row=2, column=6, value=sourceActSheet.cell_value(rowx=1, colx=13))#Voltage LG
destinationActSheet.cell(row=2, column=7, value=sourceActSheet.cell_value(rowx=2, colx=13))#Bus Rating
destinationActSheet.cell(row=2, column=8, value=sourceActSheet.cell_value(rowx=3, colx=13))#Phase/Wire
destinationActSheet.cell(row=2, column=9, value=sourceActSheet.cell_value(rowx=1, colx=15))#SC Rating
destinationActSheet.cell(row=2, column=10, value=sourceActSheet.cell_value(rowx=2, colx=18))#Mounting
destinationActSheet.cell(row=2, column=11, value=sourceActSheet.cell_value(rowx=3, colx=18))#Enclosure
destinationActSheet.cell(row=2, column=12, value="")#Emergency
destinationActSheet.cell(row=2, column=13, value="")#UPS
destinationActSheet.cell(row=2, column=14, value="")#Location DWG
destinationActSheet.cell(row=2, column=15, value="")#Circuit Number
destinationActSheet.cell(row=2, column=16, value="")#Description
destinationActSheet.cell(row=2, column=17, value="")#Leaf
destinationActSheet.cell(row=2, column=18, value="")#Load Type
destinationActSheet.cell(row=2, column=19, value="")#EA
destinationActSheet.cell(row=2, column=20, value="")#Q
destinationActSheet.cell(row=2, column=21, value="")#Dem
destinationActSheet.cell(row=2, column=22, value="")#Total VA
destinationActSheet.cell(row=2, column=23, value=sourceActSheet.cell_value(rowx=3, colx=6))#BRK
destinationActSheet.cell(row=2, column=24, value="")#PH

#Header Cleanup
# A systematic mistake is that of not declaring if a panel is fed via a transformer

#Define Panel Size
panelSize = 0
myList = sourceActSheet.col_values(panelSize, start_rowx=5, end_rowx=32)
# using remove() to perform removal
while("" in myList):
    myList.remove("")
myList = [ int(x) for x in myList ]
panelSize = math.ceil(max(myList)/2)
equipmentCount = [ int((x/2)+1) for x in myList]
#Offset for correct Print
equipmentCount = [ x+2 for x in equipmentCount]

#Populate Subsequent Panel Schedule Values (Uneven)
for items in equipmentCount:
    destinationActSheet.cell(row=items, column=1, value=sourceActSheet.cell_value(rowx=1, colx=2) + "-" + str(int(sourceActSheet.cell_value(rowx=items+2, colx=0))))#Equipment
    destinationActSheet.cell(row=items, column=2, value=sourceActSheet.cell_value(rowx=1, colx=2))#Source
    destinationActSheet.cell(row=items, column=3, value=sourceActSheet.cell_value(rowx=items+2, colx=2))#Location
    destinationActSheet.cell(row=items, column=4, value="")#Room
    destinationActSheet.cell(row=items, column=5, value="")#Voltage LL
    destinationActSheet.cell(row=items, column=6, value="")#Voltage LG
    destinationActSheet.cell(row=items, column=7, value="")#Bus Rating
    destinationActSheet.cell(row=items, column=8, value="")#Phase/Wire
    destinationActSheet.cell(row=items, column=9, value="")#SC Rating
    destinationActSheet.cell(row=items, column=10, value="")#Mounting
    destinationActSheet.cell(row=items, column=11, value="")#Enclosure
    destinationActSheet.cell(row=items, column=12, value="")#Emergency
    destinationActSheet.cell(row=items, column=13, value="")#UPS
    destinationActSheet.cell(row=items, column=14, value="")#Location DWG
    destinationActSheet.cell(row=items, column=15, value=str(int(sourceActSheet.cell_value(rowx=items+2, colx=0))))#Circuit Number
    destinationActSheet.cell(row=items, column=16, value=str(sourceActSheet.cell_value(rowx=items+2, colx=1)))#Description
    destinationActSheet.cell(row=items, column=17, value="")#Panel Schedule
    destinationActSheet.cell(row=items, column=18, value=sourceActSheet.cell_value(rowx=items+2, colx=3))#Load Type
    destinationActSheet.cell(row=items, column=19, value=str(sourceActSheet.cell_value(rowx=items+2, colx=4)))#EA
    destinationActSheet.cell(row=items, column=20, value=str(sourceActSheet.cell_value(rowx=items+2, colx=5)))#Q
    destinationActSheet.cell(row=items, column=21, value=str(sourceActSheet.cell_value(rowx=items+2, colx=6)))#Dem
    destinationActSheet.cell(row=items, column=22, value=str(sourceActSheet.cell_value(rowx=items+2, colx=7)))#Total VA
    #Clean strings
    value=str(sourceActSheet.cell_value(rowx=items+2, colx=8))
    if is_float(value):
        destinationActSheet.cell(row=items, column=23, value=str(int(sourceActSheet.cell_value(rowx=items+2, colx=8))))#BRK
    else:
        destinationActSheet.cell(row=items, column=23, value=str(sourceActSheet.cell_value(rowx=items+2, colx=8)))#BRK
    #Clean strings
    value=str(sourceActSheet.cell_value(rowx=items+2, colx=9))
    if is_float(value):
        destinationActSheet.cell(row=items, column=24, value=str(int(sourceActSheet.cell_value(rowx=items+2, colx=9))))#BRK
    else:
        destinationActSheet.cell(row=items, column=24, value=str(sourceActSheet.cell_value(rowx=items+2, colx=9)))#BRK

#Equipment Label Cleanup


#Save changes
wbFormat.save(myDestPath)